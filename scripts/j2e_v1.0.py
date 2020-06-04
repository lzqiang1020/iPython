# -*- coding: UTF-8 -*-
# @File        :  j2e_v1.0
# @CreateTime  :  2020/4/10 17:40

import json
import openpyxl

# 读取 .json 文件
def read_json(file):
    with open(file, 'r', encoding='utf8') as fr:
        json_data = json.load(fr)
    return json_data

# 将数据写入 excel 文件
def write_to_excel(file):
    j_data = read_json(file)
    j_data_length = len(j_data)

    # 创建 excel 文件并获取默认的 sheet 表
    wb = openpyxl.Workbook()
    ws = wb.active

    # 新建表头
    table_title_field = list(j_data[0].keys())
    for col in range(len(table_title_field)):
        last_col = col + 1
        ws.cell(row=1, column=last_col).value = table_title_field[col]

    # 写入基本数据
    for i in range(j_data_length):
        each_line = j_data[i]
        each_line_field = list(each_line.keys())

        for j in range(len(each_line_field)):
            ws.cell(row=i + 2, column=j + 1).value = str(each_line[each_line_field[j]])

    wb.save('je_v1.0.xlsx')

if __name__ == '__main__':
    write_to_excel('./test.json')