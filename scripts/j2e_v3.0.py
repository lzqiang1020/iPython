# -*- coding: UTF-8 -*-
# @File        :  json_to_exc_tem
# @CreateTime  :  2020/4/10 17:40

import json
import openpyxl

# 读取 .json 文件
# def read_json(file):
#     with open(file, 'r', encoding='utf8') as fr:
#         json_data = json.load(fr)
#     return json_data

# 将数据写入 excel 文件
def write_to_excel(file):
    # j_data = read_json(file)
    j_data = [
        {
            "id": 34,
            "other": ["str_1", "str_2", "str_3"]
        },
        {
            "id": 24,
            "name": "结构化（SDS）",
            "other": ["str_1", "str_2", "str_4",]
        },
        {
            "id": 25,
            "name": "管理",
        }
    ]
    j_data_length = len(j_data)
    # 创建 excel 文件并获取默认的 sheet 表
    wb = openpyxl.Workbook()
    ws = wb.active
    # sheet_je = openpyxl.workbook().create_sheet('je')   # 创建 excel 文件并在默认的 sheet 表后追加一个sheet表‘je’，在默认表前建sheet：openpyxl.workbook().create_sheet('je', 0)

    # 新建表头
    table_title_field = list(j_data[0].keys())
    last_col = 0
    for col in range(len(table_title_field)):
        last_col = col + 1
        ws.cell(row=1, column=last_col).value = table_title_field[col]

    # 写入基本数据
    for i in range(j_data_length):
        each_line = j_data[i]
        each_line_field = list(each_line.keys())
        c = 1
        if i == 0:
            for c in range(len(each_line_field)):
                ws.cell(row=1, column=c + 1).value = each_line_field[c]


        j = 0
        while j < len(each_line_field):
            if each_line_field[j] not in table_title_field:
                table_title_field.append(each_line_field[j])
                last_col += 1
                ws.cell(row=1, column=last_col).value = each_line_field[j]

            column_f = table_title_field.index(each_line_field[j]) + 1
            ws.cell(row=i+2, column=column_f).value = str(each_line[each_line_field[j]])
            j += 1

    wb.save('je_v3.0.xlsx')

if __name__ == '__main__':
    # write_to_excel('./test.json')
    write_to_excel()